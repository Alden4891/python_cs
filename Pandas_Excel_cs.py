# Pandas Excel Cheat Sheet

# ------------------------------
# 1. Installation
# ------------------------------

# Install Pandas and Openpyxl:
# Comment: Install Pandas library and Openpyxl for Excel file handling.
# Syntax: pip install pandas openpyxl
pip install pandas openpyxl

# ------------------------------
# 2. Importing Libraries
# ------------------------------

# Import Pandas:
# Comment: Import the pandas library for data manipulation.
import pandas as pd

# ------------------------------
# 3. Reading Excel Files
# ------------------------------

# Read Excel File:
# Comment: Read an Excel file into a DataFrame.
# Syntax: pd.read_excel(file_path, sheet_name=None)
df = pd.read_excel("file.xlsx")

# Read Specific Sheet:
# Comment: Read a specific sheet from an Excel file.
# Syntax: pd.read_excel(file_path, sheet_name="Sheet1")
df = pd.read_excel("file.xlsx", sheet_name="Sheet1")

# Read Multiple Sheets:
# Comment: Read multiple sheets from an Excel file into a dictionary of DataFrames.
# Syntax: pd.read_excel(file_path, sheet_name=["Sheet1", "Sheet2"])
dfs = pd.read_excel("file.xlsx", sheet_name=["Sheet1", "Sheet2"])

# ------------------------------
# 4. Writing Excel Files
# ------------------------------

# Write DataFrame to Excel:
# Comment: Write a DataFrame to an Excel file.
# Syntax: df.to_excel(file_path, sheet_name="Sheet1", index=False)
df.to_excel("output.xlsx", sheet_name="Sheet1", index=False)

# Write Multiple DataFrames:
# Comment: Write multiple DataFrames to different sheets in the same Excel file.
# Syntax: pd.ExcelWriter(file_path) with writer
with pd.ExcelWriter("output.xlsx") as writer:
    df1.to_excel(writer, sheet_name="Sheet1", index=False)
    df2.to_excel(writer, sheet_name="Sheet2", index=False)

# ------------------------------
# 5. Reading and Writing Specific Columns
# ------------------------------

# Read Specific Columns:
# Comment: Read specific columns from an Excel file.
# Syntax: pd.read_excel(file_path, usecols=["Column1", "Column2"])
df = pd.read_excel("file.xlsx", usecols=["Column1", "Column2"])

# Write Specific Columns:
# Comment: Write only specific columns to an Excel file.
# Syntax: df.to_excel(file_path, columns=["Column1", "Column2"], index=False)
df.to_excel("output.xlsx", columns=["Column1", "Column2"], index=False)

# ------------------------------
# 6. Handling Excel File Formats
# ------------------------------

# Read Excel File with Different Engine:
# Comment: Specify the engine to read Excel files (e.g., "openpyxl", "xlrd").
# Syntax: pd.read_excel(file_path, engine="openpyxl")
df = pd.read_excel("file.xlsx", engine="openpyxl")

# ------------------------------
# 7. Working with Excel Styles
# ------------------------------

# Apply Styles in Excel:
# Comment: Apply styles to cells in an Excel file.
# Syntax: Use ExcelWriter with formatting options
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Create a new Excel file with styling
with pd.ExcelWriter("styled_output.xlsx", engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name="StyledSheet", index=False)
    workbook = writer.book
    worksheet = workbook["StyledSheet"]
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in worksheet["A"]:
        cell.fill = fill

# ------------------------------
# 8. Handling Missing Data
# ------------------------------

# Read Excel File and Handle Missing Data:
# Comment: Handle missing data during read.
# Syntax: pd.read_excel(file_path, na_values=["NA", "N/A"])
df = pd.read_excel("file.xlsx", na_values=["NA", "N/A"])

# Write DataFrame and Handle Missing Data:
# Comment: Fill missing data before writing to an Excel file.
# Syntax: df.fillna(value="default_value").to_excel(file_path, index=False)
df.fillna(value="Unknown").to_excel("output_filled.xlsx", index=False)

# ------------------------------
# 9. Filtering and Manipulating Data
# ------------------------------

# Filter DataFrame Based on Condition:
# Comment: Filter rows based on condition before writing to Excel.
# Syntax: df[df["Column"] > value]
filtered_df = df[df["Column"] > 50]

# ------------------------------
# 10. Saving to Excel with Conditional Formatting
# ------------------------------

# Save DataFrame with Conditional Formatting:
# Comment: Save DataFrame with conditional formatting using openpyxl.
from openpyxl.formatting.rule import ColorScaleRule

# Write DataFrame to Excel
with pd.ExcelWriter("conditional_formatting.xlsx", engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name="FormattedSheet", index=False)
    workbook = writer.book
    worksheet = workbook["FormattedSheet"]

    # Apply conditional formatting
    rule = ColorScaleRule(start_type='min', start_color='FF0000',
                          end_type='max', end_color='00FF00')
    worksheet.conditional_formatting.add('B2:B10', rule)
